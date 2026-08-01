from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from portrait_bot.access_codes import (
    access_code_report_rows,
    access_code_stats,
    build_access_codes_workbook,
    create_access_code_batch,
    extract_access_code,
    redeem_access_code,
)
from portrait_bot.services import (
    add_ledger_entry,
    add_wallet_entry,
    balance,
    create_generation,
    get_or_create_user,
    wallet_balance,
)


def test_extract_access_code_from_copied_text() -> None:
    assert extract_access_code("Код:  YAI-2345-ABCD\n") == "YAI-2345-ABCD"
    assert extract_access_code("yai 2345 abcd") == "YAI-2345-ABCD"
    assert extract_access_code("без кода") is None


async def test_access_code_can_only_be_redeemed_once(database) -> None:
    async with database.sessions() as session:
        batch, codes = await create_access_code_batch(
            session,
            count=3,
            accesses_per_code=5,
            created_by=999,
            expires_in_days=30,
        )
        assert batch.code_count == 3
        assert len({code.code for code in codes}) == 3

        user = await get_or_create_user(session, telegram_id=10001, username="buyer")
        redeemed = await redeem_access_code(session, user=user, raw_code=codes[0].code.lower())
        assert redeemed.redeemed_by_telegram_id == 10001
        assert await balance(session, user.id) == 5

        with pytest.raises(ValueError, match="access_code_redeemed"):
            await redeem_access_code(session, user=user, raw_code=codes[0].code)

        stats = await access_code_stats(session)
        assert stats.total == 3
        assert stats.active == 2
        assert stats.redeemed == 1


async def test_persisted_expiring_access_code_can_be_redeemed(database) -> None:
    async with database.sessions() as session:
        _, codes = await create_access_code_batch(
            session,
            count=1,
            accesses_per_code=3,
            created_by=999,
            expires_in_days=30,
        )
        raw_code = codes[0].code
        user = await get_or_create_user(session, telegram_id=10004, username="persisted")

    # A fresh SQLite session loads DateTime values without timezone metadata,
    # matching the production path that previously raised TypeError.
    async with database.sessions() as session:
        user = await get_or_create_user(session, telegram_id=10004, username="persisted")
        redeemed = await redeem_access_code(session, user=user, raw_code=raw_code)
        assert redeemed.redeemed_by_telegram_id == 10004
        assert await balance(session, user.id) == 3


async def test_access_code_report_is_a_valid_excel_file(database) -> None:
    async with database.sessions() as session:
        await create_access_code_batch(
            session,
            count=2,
            accesses_per_code=10,
            created_by=999,
            expires_in_days=None,
            title="Тестовая акция",
        )
        stats = await access_code_stats(session)
        rows = await access_code_report_rows(session)
    content = build_access_codes_workbook(rows, stats)
    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == ["Сводка", "Коды"]
    assert workbook["Коды"].max_row == 3
    assert workbook["Коды"]["D2"].value == "Тестовая акция"


async def test_generation_uses_access_before_ruble_balance(database, settings) -> None:
    async with database.sessions() as session:
        user = await get_or_create_user(session, telegram_id=10002)
        await add_ledger_entry(
            session,
            user_id=user.id,
            amount=2,
            entry_type="test_access",
            idempotency_key="test-access-2",
        )
        await add_wallet_entry(
            session,
            user_id=user.id,
            amount_rub=Decimal("500"),
            entry_type="test_wallet",
            idempotency_key="test-wallet-500",
        )
        generation = await create_generation(
            session,
            settings,
            user=user,
            source=b"photo",
            prompt="test",
            mode="photo:enhance",
            quantity=1,
            price_rub=Decimal("99"),
        )
        assert generation.credits == 1
        assert Decimal(generation.price_rub) == Decimal("0")
        assert await balance(session, user.id) == 1
        assert await wallet_balance(session, user.id) == Decimal("500")


async def test_video_keeps_using_ruble_balance(database, settings) -> None:
    async with database.sessions() as session:
        user = await get_or_create_user(session, telegram_id=10003)
        await add_ledger_entry(
            session,
            user_id=user.id,
            amount=2,
            entry_type="test_access",
            idempotency_key="test-video-access-2",
        )
        await add_wallet_entry(
            session,
            user_id=user.id,
            amount_rub=Decimal("500"),
            entry_type="test_wallet",
            idempotency_key="test-video-wallet-500",
        )
        generation = await create_generation(
            session,
            settings,
            user=user,
            source=b"photo",
            prompt="test",
            mode="video:5",
            quantity=1,
            price_rub=Decimal("200"),
        )
        assert generation.credits == 0
        assert Decimal(generation.price_rub) == Decimal("200")
        assert await balance(session, user.id) == 2
        assert await wallet_balance(session, user.id) == Decimal("300")
