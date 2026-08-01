from __future__ import annotations

import re
import secrets
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from portrait_bot.models import AccessCode, AccessCodeBatch, LedgerEntry, User

CODE_ALPHABET = "23456789ABCDEFGHJKLMNPQRSTUVWXYZ"
ACCESS_CODE_PATTERN = re.compile(
    r"(?i)(?<![A-Z0-9])YAI[\s-]*([2-9A-HJ-NP-Z]{4})[\s-]*([2-9A-HJ-NP-Z]{4})(?![A-Z0-9])"
)


@dataclass(slots=True, frozen=True)
class AccessCodeStats:
    total: int
    active: int
    redeemed: int
    expired: int
    disabled: int


def normalize_access_code(raw: str) -> str:
    compact = "".join(character for character in raw.upper() if character.isalnum())
    if compact.startswith("YAI"):
        compact = compact[3:]
    return f"YAI-{compact[:4]}-{compact[4:8]}" if len(compact) == 8 else raw.strip().upper()


def extract_access_code(raw: str) -> str | None:
    """Return the first access code found in copied text or a TXT report."""
    match = ACCESS_CODE_PATTERN.search(raw)
    if match is None:
        return None
    return f"YAI-{match.group(1).upper()}-{match.group(2).upper()}"


def _new_code() -> str:
    body = "".join(secrets.choice(CODE_ALPHABET) for _ in range(8))
    return f"YAI-{body[:4]}-{body[4:]}"


def _aware(value: datetime) -> datetime:
    return value.replace(tzinfo=UTC) if value.tzinfo is None else value


def access_code_display_status(code: AccessCode, *, now: datetime | None = None) -> str:
    if code.status == "redeemed" or code.redeemed_at is not None:
        return "redeemed"
    if code.status == "disabled":
        return "disabled"
    current = now or datetime.now(UTC)
    if code.expires_at is not None and _aware(code.expires_at) <= current:
        return "expired"
    return "active"


async def create_access_code_batch(
    session: AsyncSession,
    *,
    count: int,
    accesses_per_code: int,
    created_by: int,
    expires_in_days: int | None,
    title: str | None = None,
) -> tuple[AccessCodeBatch, list[AccessCode]]:
    if not 1 <= count <= 500:
        raise ValueError("invalid_code_count")
    if not 1 <= accesses_per_code <= 1000:
        raise ValueError("invalid_access_count")
    if expires_in_days is not None and not 1 <= expires_in_days <= 3650:
        raise ValueError("invalid_expiry")

    now = datetime.now(UTC)
    expires_at = now + timedelta(days=expires_in_days) if expires_in_days else None
    batch = AccessCodeBatch(
        title=(title or f"Партия {now:%d.%m.%Y %H:%M}"),
        code_count=count,
        accesses_per_code=accesses_per_code,
        created_by=created_by,
        created_at=now,
        expires_at=expires_at,
    )
    session.add(batch)
    await session.flush()

    existing = set((await session.scalars(select(AccessCode.code))).all())
    generated: set[str] = set()
    while len(generated) < count:
        candidate = _new_code()
        if candidate not in existing:
            generated.add(candidate)
    codes = [
        AccessCode(
            batch_id=batch.id,
            code=code,
            accesses=accesses_per_code,
            status="active",
            created_at=now,
            expires_at=expires_at,
        )
        for code in sorted(generated)
    ]
    session.add_all(codes)
    await session.commit()
    return batch, codes


async def redeem_access_code(
    session: AsyncSession,
    *,
    user: User,
    raw_code: str,
) -> AccessCode:
    normalized = normalize_access_code(raw_code)
    code = await session.scalar(select(AccessCode).where(AccessCode.code == normalized))
    if code is None:
        raise ValueError("access_code_not_found")
    status = access_code_display_status(code)
    if status == "redeemed":
        raise ValueError("access_code_redeemed")
    if status == "expired":
        raise ValueError("access_code_expired")
    if status != "active":
        raise ValueError("access_code_disabled")

    now = datetime.now(UTC)
    result = await session.execute(
        update(AccessCode)
        .where(
            AccessCode.id == code.id,
            AccessCode.status == "active",
            AccessCode.redeemed_at.is_(None),
            or_(AccessCode.expires_at.is_(None), AccessCode.expires_at > now),
        )
        .values(
            status="redeemed",
            redeemed_by_user_id=user.id,
            redeemed_by_telegram_id=user.telegram_id,
            redeemed_at=now,
        )
        # SQLite returns persisted DateTime values without timezone data.  The
        # ORM's default in-memory synchronization would compare that naive
        # value with ``now`` (UTC-aware) and fail before executing the UPDATE.
        # The row is refreshed explicitly below, so no synchronization is
        # needed here.
        .execution_options(synchronize_session=False)
    )
    if result.rowcount != 1:
        await session.rollback()
        raise ValueError("access_code_redeemed")
    session.add(
        LedgerEntry(
            user_id=user.id,
            amount=code.accesses,
            entry_type="access_code",
            idempotency_key=f"access-code:{code.id}",
            reference_type="access_code",
            reference_id=code.id,
            comment=f"Код {code.code}",
            expires_at=code.expires_at,
        )
    )
    await session.commit()
    await session.refresh(code)
    return code


async def access_code_stats(session: AsyncSession) -> AccessCodeStats:
    now = datetime.now(UTC)
    total = int(await session.scalar(select(func.count(AccessCode.id))) or 0)
    redeemed = int(
        await session.scalar(
            select(func.count(AccessCode.id)).where(AccessCode.status == "redeemed")
        )
        or 0
    )
    disabled = int(
        await session.scalar(
            select(func.count(AccessCode.id)).where(AccessCode.status == "disabled")
        )
        or 0
    )
    expired = int(
        await session.scalar(
            select(func.count(AccessCode.id)).where(
                AccessCode.status == "active",
                AccessCode.expires_at.is_not(None),
                AccessCode.expires_at <= now,
            )
        )
        or 0
    )
    active = max(total - redeemed - disabled - expired, 0)
    return AccessCodeStats(
        total=total,
        active=active,
        redeemed=redeemed,
        expired=expired,
        disabled=disabled,
    )


async def recent_access_codes(
    session: AsyncSession,
    *,
    status: str,
    limit: int = 20,
) -> list[AccessCode]:
    now = datetime.now(UTC)
    query = select(AccessCode)
    if status == "redeemed":
        query = query.where(AccessCode.status == "redeemed").order_by(
            AccessCode.redeemed_at.desc()
        )
    elif status == "active":
        query = query.where(
            AccessCode.status == "active",
            or_(AccessCode.expires_at.is_(None), AccessCode.expires_at > now),
        ).order_by(AccessCode.created_at.desc())
    else:
        query = query.order_by(AccessCode.created_at.desc())
    return list((await session.scalars(query.limit(limit))).all())


async def access_code_report_rows(
    session: AsyncSession,
) -> list[tuple[AccessCode, str, User | None]]:
    result = await session.execute(
        select(AccessCode, AccessCodeBatch.title, User)
        .join(AccessCodeBatch, AccessCode.batch_id == AccessCodeBatch.id)
        .outerjoin(User, AccessCode.redeemed_by_user_id == User.id)
        .order_by(AccessCode.created_at.desc())
    )
    return list(result.all())


def build_access_codes_workbook(
    rows: list[tuple[AccessCode, str, User | None]],
    stats: AccessCodeStats,
) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Сводка"
    summary.append(["Отчёт по кодам доступа"])
    summary.append(["Сформирован", datetime.now(UTC).strftime("%d.%m.%Y %H:%M UTC")])
    summary.append([])
    summary.append(["Показатель", "Количество"])
    summary.append(["Всего кодов", stats.total])
    summary.append(["Оставшиеся", stats.active])
    summary.append(["Использованные", stats.redeemed])
    summary.append(["Истёкшие", stats.expired])
    summary.append(["Отключённые", stats.disabled])
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 24

    sheet = workbook.create_sheet("Коды")
    headers = [
        "Код",
        "Статус",
        "Доступов",
        "Партия",
        "Создан",
        "Действует до",
        "Telegram ID",
        "Username",
        "Активирован",
    ]
    sheet.append(headers)
    status_labels = {
        "active": "Оставшийся",
        "redeemed": "Использован",
        "expired": "Истёк",
        "disabled": "Отключён",
    }
    for code, batch_title, user in rows:
        status = access_code_display_status(code)
        sheet.append(
            [
                code.code,
                status_labels[status],
                code.accesses,
                batch_title,
                _aware(code.created_at).strftime("%d.%m.%Y %H:%M"),
                _aware(code.expires_at).strftime("%d.%m.%Y %H:%M")
                if code.expires_at
                else "Без срока",
                code.redeemed_by_telegram_id or "",
                f"@{user.username}" if user and user.username else "",
                _aware(code.redeemed_at).strftime("%d.%m.%Y %H:%M")
                if code.redeemed_at
                else "",
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for target in (summary, sheet):
        target.freeze_panes = "A2" if target is sheet else "A5"
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for cell in summary[1]:
        cell.font = Font(size=16, bold=True, color="1F4E78")
    for cell in summary[4]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.auto_filter.ref = sheet.dimensions
    widths = [20, 16, 12, 30, 20, 20, 18, 24, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
